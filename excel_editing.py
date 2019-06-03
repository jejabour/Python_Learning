import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)

    sheet = wb['Sheet1']

    # A way of specifying a cell
    cell = sheet['a1']

    # Same as above
    # cell = sheet.cell(1, 1)

    # Prints value of a cell
    # print(cell.value)

    # Prints the number of rows in the spreadsheet
    # print(sheet.max_row)

    # We do +1 because otherwise it would only go to but not include the max
    #If we start at 2, we can skip the headers
    for row in range(2, sheet.max_row + 1): 

        # This will get the third column
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price = round(corrected_price, 2)
        
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

        # Never figured out how to actually change the format of the cell to 'Currency'
        # But this is the closest I got
        corrected_price_cell.number_format = "$###0.00"

    # Sets a reference point for use in the chart
    values = Reference(sheet,
            min_row = 2,
            max_row = sheet.max_row,
            min_col = 4,
            max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    wb.save(filename)
